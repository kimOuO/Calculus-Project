"""
Students Actor - 學生資料管理
"""
import json
import logging
import io
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.db import transaction
from django.http import HttpResponse
try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import PatternFill, Font
except ImportError:
    load_workbook = None
    Workbook = None
    PatternFill = None
    Font = None

from main.apps.Calculus_metadata.models import Students, Score
from main.apps.Calculus_metadata.serializers import StudentsWriteSerializer, StudentsReadSerializer
from main.apps.Calculus_metadata.services.common import UuidService, TimestampService, ValidationService
from main.apps.Calculus_metadata.services.business import SqlDbBusinessService
from main.utils.response import success_response, error_response

logger = logging.getLogger(__name__)


class StudentActor:
    """學生資料 Actor - 處理學生相關的所有業務操作"""
    
    @staticmethod
    @csrf_exempt
    @require_http_methods(["POST"])
    @transaction.atomic
    def create(request):
        """
        創建學生
        POST /api/v0.1/Calculus_oom/Calculus_metadata/Student_MetadataWriter/create
        """
        try:
            # Step 1: 解析請求
            data = json.loads(request.body)
            logger.info(f"Creating student with data: {data}")
            
            # Step 2: 驗證數據
            serializer = StudentsWriteSerializer(data=data)
            if not serializer.is_valid():
                return error_response("Validation failed", serializer.errors, 400)
            
            validated_data = serializer.validated_data
            
            # Step 3: 生成 UUID 和時間戳
            student_uuid = UuidService.generate_student_uuid(validated_data['student_semester'])
            timestamp = TimestampService.get_current_timestamp()
            
            # Step 4: 準備完整數據
            complete_data = {
                'student_uuid': student_uuid,
                'student_name': validated_data['student_name'],
                'student_number': validated_data['student_number'],
                'student_semester': validated_data['student_semester'],
                'student_status': validated_data.get('student_status', '修業中'),
                'student_created_at': timestamp,
                'student_updated_at': timestamp,
            }
            
            # Step 5: 創建學生（通過 Business Service）
            student = SqlDbBusinessService.create_entity(Students, complete_data)
            
            # Step 6: 同時創建對應的成績記錄
            score_uuid = UuidService.generate_score_uuid(validated_data['student_semester'])
            score_data = {
                'score_uuid': score_uuid,
                'f_student_uuid': student_uuid,
                'score_quiz1': '',
                'score_midterm': '',
                'score_quiz2': '',
                'score_finalexam': '',
                'score_total': '',
                'score_created_at': timestamp,
                'score_updated_at': timestamp,
            }
            SqlDbBusinessService.create_entity(Score, score_data)
            
            # Step 7: 格式化輸出
            output = StudentsReadSerializer(student).data
            logger.info(f"Student created successfully: {student_uuid}")
            
            return success_response(output, "Student created successfully", 201)
            
        except json.JSONDecodeError:
            return error_response("Invalid JSON format", None, 400)
        except Exception as e:
            logger.error(f"Error creating student: {str(e)}")
            return error_response(f"Unknown error: {str(e)}", None, 500)
    
    @staticmethod
    @csrf_exempt
    @require_http_methods(["POST"])
    def read(request):
        """
        查詢學生
        POST /api/v0.1/Calculus_oom/Calculus_metadata/Student_MetadataWriter/read
        """
        try:
            # Step 1: 解析請求
            data = json.loads(request.body)
            logger.info(f"Reading students with filters: {data}")
            
            # Step 2: 判斷查詢類型
            if 'student_uuid' in data:
                # 單個查詢
                student = SqlDbBusinessService.get_entity(Students, 'student_uuid', data['student_uuid'])
                if not student:
                    return error_response("Student not found", None, 404)
                output = StudentsReadSerializer(student).data
            elif data:
                # 條件查詢
                students = SqlDbBusinessService.get_entities(Students, data)
                output = StudentsReadSerializer(students, many=True).data
            else:
                # 查詢全部
                students = SqlDbBusinessService.get_entities(Students, {})
                output = StudentsReadSerializer(students, many=True).data
            
            logger.info(f"Students retrieved successfully")
            return success_response(output, "Students retrieved successfully", 200)
            
        except json.JSONDecodeError:
            return error_response("Invalid JSON format", None, 400)
        except Exception as e:
            logger.error(f"Error reading students: {str(e)}")
            return error_response(f"Unknown error: {str(e)}", None, 500)
    
    @staticmethod
    @csrf_exempt
    @require_http_methods(["POST"])
    @transaction.atomic
    def update(request):
        """
        更新學生
        POST /api/v0.1/Calculus_oom/Calculus_metadata/Student_MetadataWriter/update
        """
        try:
            # Step 1: 解析請求
            data = json.loads(request.body)
            logger.info(f"Updating student with data: {data}")
            
            # Step 2: 驗證必要欄位
            is_valid, missing_keys = ValidationService.validate_required_keys(data, ['student_uuid'])
            if not is_valid:
                return error_response(f"Missing required keys: {missing_keys}", None, 400)
            
            # Step 3: 查詢學生
            student = SqlDbBusinessService.get_entity(Students, 'student_uuid', data['student_uuid'])
            if not student:
                return error_response("Student not found", None, 404)
            
            # Step 4: 驗證更新數據
            update_fields = {k: v for k, v in data.items() if k != 'student_uuid'}
            serializer = StudentsWriteSerializer(data=update_fields, partial=True)
            if not serializer.is_valid():
                return error_response("Validation failed", serializer.errors, 400)
            
            # Step 5: 更新時間戳
            update_fields['student_updated_at'] = TimestampService.get_current_timestamp()
            
            # Step 6: 執行更新
            updated_student = SqlDbBusinessService.update_entity(student, update_fields)
            
            # Step 7: 格式化輸出
            output = StudentsReadSerializer(updated_student).data
            logger.info(f"Student updated successfully: {data['student_uuid']}")
            
            return success_response(output, "Student updated successfully", 200)
            
        except json.JSONDecodeError:
            return error_response("Invalid JSON format", None, 400)
        except Exception as e:
            logger.error(f"Error updating student: {str(e)}")
            return error_response(f"Unknown error: {str(e)}", None, 500)
    
    @staticmethod
    @csrf_exempt
    @require_http_methods(["POST"])
    @transaction.atomic
    def delete(request):
        """
        刪除學生（級聯刪除相關成績）
        POST /api/v0.1/Calculus_oom/Calculus_metadata/Student_MetadataWriter/delete
        """
        try:
            # Step 1: 解析請求
            data = json.loads(request.body)
            logger.info(f"Deleting student with data: {data}")
            
            # Step 2: 驗證必要欄位
            is_valid, missing_keys = ValidationService.validate_required_keys(data, ['student_uuid'])
            if not is_valid:
                return error_response(f"Missing required keys: {missing_keys}", None, 400)
            
            # Step 3: 查詢學生
            student = SqlDbBusinessService.get_entity(Students, 'student_uuid', data['student_uuid'])
            if not student:
                return error_response("Student not found", None, 404)
            
            # Step 4: 刪除相關成績（級聯刪除）
            deleted_scores = SqlDbBusinessService.delete_entities(Score, {'f_student_uuid': data['student_uuid']})
            logger.info(f"Deleted {deleted_scores} related scores")
            
            # Step 5: 刪除學生
            SqlDbBusinessService.delete_entity(student)
            
            logger.info(f"Student deleted successfully: {data['student_uuid']}")
            return success_response(None, "Student and related scores deleted successfully", 200)
            
        except json.JSONDecodeError:
            return error_response("Invalid JSON format", None, 400)
        except Exception as e:
            logger.error(f"Error deleting student: {str(e)}")
            return error_response(f"Unknown error: {str(e)}", None, 500)
    
    @staticmethod
    @csrf_exempt
    @require_http_methods(["POST"])
    @transaction.atomic
    def status(request):
        """
        更新學生狀態
        POST /api/v0.1/Calculus_oom/Calculus_metadata/Student_MetadataWriter/status
        """
        try:
            # Step 1: 解析請求
            data = json.loads(request.body)
            logger.info(f"Updating student status with data: {data}")
            
            # Step 2: 驗證必要欄位
            is_valid, missing_keys = ValidationService.validate_required_keys(data, ['student_uuid', 'student_status'])
            if not is_valid:
                return error_response(f"Missing required keys: {missing_keys}", None, 400)
            
            # Step 3: 驗證狀態值
            allowed_statuses = ["修業中", "二退", "被當", "修業完畢"]
            if data['student_status'] not in allowed_statuses:
                return error_response(f"Invalid status. Must be one of: {', '.join(allowed_statuses)}", None, 400)
            
            # Step 4: 查詢學生
            student = SqlDbBusinessService.get_entity(Students, 'student_uuid', data['student_uuid'])
            if not student:
                return error_response("Student not found", None, 404)
            
            # Step 5: 更新狀態
            update_data = {
                'student_status': data['student_status'],
                'student_updated_at': TimestampService.get_current_timestamp()
            }
            
            # Step 6: 如果狀態改為「二退」，清空該學生的所有成績
            if data['student_status'] == '二退':
                scores = SqlDbBusinessService.get_entities(Score, {'f_student_uuid': data['student_uuid']})
                for score in scores:
                    clear_data = {
                        'score_quiz1': '',
                        'score_midterm': '',
                        'score_quiz2': '',
                        'score_finalexam': '',
                        'score_total': '',
                        'score_updated_at': TimestampService.get_current_timestamp()
                    }
                    SqlDbBusinessService.update_entity(score, clear_data)
                logger.info(f"Cleared scores for student: {data['student_uuid']}")
            
            updated_student = SqlDbBusinessService.update_entity(student, update_data)
            
            # Step 7: 格式化輸出
            output = StudentsReadSerializer(updated_student).data
            logger.info(f"Student status updated successfully: {data['student_uuid']}")
            
            return success_response(output, "Student status updated successfully", 200)
            
        except json.JSONDecodeError:
            return error_response("Invalid JSON format", None, 400)
        except Exception as e:
            logger.error(f"Error updating student status: {str(e)}")
            return error_response(f"Unknown error: {str(e)}", None, 500)
    
    @staticmethod
    @csrf_exempt
    @require_http_methods(["POST"])
    @transaction.atomic
    def upload_excel(request):
        """
        批量上傳學生資料 (Excel)
        POST /api/v0.1/Calculus_oom/Calculus_metadata/Student_MetadataWriter/upload_excel
        """
        try:
            # Step 1: 檢查 openpyxl 是否安裝
            if load_workbook is None or Workbook is None:
                return error_response(
                    "Excel support not available. Please install openpyxl: pip install openpyxl",
                    None,
                    500
                )
            
            # Step 2: 解析上傳檔案與學期參數
            uploaded_file = request.FILES.get('file')
            if not uploaded_file:
                return error_response("No file uploaded", None, 400)

            student_semester = request.POST.get('student_semester', '').strip()
            if not student_semester:
                return error_response("Missing required field: student_semester", None, 400)

            logger.info(f"Uploading student Excel file: {uploaded_file.name}, semester: {student_semester}")

            # Step 3: 讀取 Excel
            try:
                workbook = load_workbook(filename=io.BytesIO(uploaded_file.read()))
                sheet = workbook.active
            except Exception as e:
                return error_response(f"Invalid Excel file: {str(e)}", None, 400)

            # Step 4: 解析資料（第一行為標題，從第二行開始）
            # Excel 欄位格式（對應 student.xlsx）:
            #   Col A (idx 0): 名字 → student_name
            #   Col B (idx 1): 姓氏 → 跳過
            #   Col C (idx 2): 學號 → student_number
            #   Col D (idx 3): 電子郵件信箱 → student_email
            #   Col E (idx 4): 科系 → 跳過
            #   Col F (idx 5): 分組 → 跳過
            created_students = []
            errors = []

            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not row or not any(row):  # 跳過空行
                    continue

                try:
                    if len(row) < 4:
                        errors.append(f"Row {row_idx}: 欄位不足（需要至少 4 欄）")
                        continue

                    student_name = str(row[0]).strip() if row[0] else None
                    student_number = str(row[2]).strip() if row[2] else None
                    student_email = str(row[3]).strip() if row[3] else ''

                    if not student_name or not student_number:
                        errors.append(f"Row {row_idx}: 名字或學號為空")
                        continue

                    # 生成 UUID 和時間戳
                    student_uuid = UuidService.generate_student_uuid(student_semester)
                    timestamp = TimestampService.get_current_timestamp()

                    # 準備完整數據
                    complete_data = {
                        'student_uuid': student_uuid,
                        'student_name': student_name,
                        'student_number': student_number,
                        'student_semester': student_semester,
                        'student_email': student_email,
                        'student_status': '修業中',
                        'student_created_at': timestamp,
                        'student_updated_at': timestamp,
                    }

                    # 創建學生（若學號已存在則跳過）
                    if Students.objects.filter(student_number=student_number).exists():
                        errors.append(f"Row {row_idx}: 學號 {student_number} 已存在，跳過")
                        continue

                    student = SqlDbBusinessService.create_entity(Students, complete_data)

                    # 創建對應的成績記錄
                    score_uuid = UuidService.generate_score_uuid(student_semester)
                    score_data = {
                        'score_uuid': score_uuid,
                        'f_student_uuid': student_uuid,
                        'score_quiz1': '',
                        'score_midterm': '',
                        'score_quiz2': '',
                        'score_finalexam': '',
                        'score_total': '',
                        'score_created_at': timestamp,
                        'score_updated_at': timestamp,
                    }
                    SqlDbBusinessService.create_entity(Score, score_data)

                    created_students.append(student_uuid)

                except Exception as e:
                    errors.append(f"Row {row_idx}: {str(e)}")
                    continue
            
            # Step 5: 格式化輸出
            output = {
                'created_count': len(created_students),
                'error_count': len(errors),
                'created_students': created_students,
                'errors': errors[:10] if errors else []  # 最多返回前 10 個錯誤
            }
            
            logger.info(f"Excel upload completed: {len(created_students)} created, {len(errors)} errors")
            
            if len(created_students) > 0:
                return success_response(
                    output,
                    f"Successfully created {len(created_students)} students with {len(errors)} errors",
                    201
                )
            else:
                return error_response("No students created", output, 400)
            
        except Exception as e:
            logger.error(f"Error uploading Excel: {str(e)}")
            return error_response(f"Unknown error: {str(e)}", None, 500)
    
    @staticmethod
    @csrf_exempt
    @require_http_methods(["POST"])
    def feedback_excel(request):
        """
        匯出學生成績 (Excel)
        POST /api/v0.1/Calculus_oom/Calculus_metadata/Student_MetadataWriter/feedback_excel
        """
        try:
            # Step 1: 檢查 openpyxl 是否安裝
            if Workbook is None:
                return error_response(
                    "Excel support not available. Please install openpyxl: pip install openpyxl",
                    None,
                    500
                )
            
            # Step 2: 解析請求
            data = json.loads(request.body)
            logger.info(f"Exporting student scores to Excel: {data}")
            
            # Step 3: 驗證必要欄位
            is_valid, missing_keys = ValidationService.validate_required_keys(data, ['student_semester'])
            if not is_valid:
                return error_response(f"Missing required keys: {missing_keys}", None, 400)
            
            semester = data['student_semester']
            
            # Step 4: 查詢該學期所有學生
            students = SqlDbBusinessService.get_entities(Students, {'student_semester': semester})
            
            if not students:
                return error_response(f"No students found for semester {semester}", None, 404)
            
            # Step 5: 創建 Excel 工作簿
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = f"成績_{semester}"

            # 設定標題行
            headers = [
                '學生名字', '學號', '第一次小考', '期中考', '第二次小考', '期末考', '最後成績', '是否被當'
            ]
            sheet.append(headers)

            # 標題行粗體
            for cell in sheet[1]:
                cell.font = Font(bold=True)

            # 紅色填充（用於被當學生）
            red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
            red_font = Font(color='CC0000', bold=True)

            # Step 6: 填充資料
            for student in students:
                score = SqlDbBusinessService.get_entity(Score, 'f_student_uuid', student.student_uuid)
                is_failed = (student.student_status == '被當')
                pass_fail_label = '被當' if is_failed else '通過'

                row_data = [
                    student.student_name,
                    student.student_number,
                    score.score_quiz1 if score else '',
                    score.score_midterm if score else '',
                    score.score_quiz2 if score else '',
                    score.score_finalexam if score else '',
                    score.score_total if score else '',
                    pass_fail_label,
                ]
                sheet.append(row_data)

                # 對被當學生整行標紅
                if is_failed:
                    current_row = sheet.max_row
                    for col_idx in range(1, len(headers) + 1):
                        cell = sheet.cell(row=current_row, column=col_idx)
                        cell.fill = red_fill
                    # 「是否被當」欄使用紅色粗體字
                    sheet.cell(row=current_row, column=len(headers)).font = red_font

            # Step 7: 生成檔案並返回
            output_stream = io.BytesIO()
            workbook.save(output_stream)
            output_stream.seek(0)
            
            response = HttpResponse(
                output_stream.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="students_scores_{semester}.xlsx"'
            
            logger.info(f"Excel exported successfully for semester {semester}")
            return response
            
        except json.JSONDecodeError:
            return error_response("Invalid JSON format", None, 400)
        except Exception as e:
            logger.error(f"Error exporting Excel: {str(e)}")
            return error_response(f"Unknown error: {str(e)}", None, 500)
